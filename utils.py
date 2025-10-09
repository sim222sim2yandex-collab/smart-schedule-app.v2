import pandas as pd
import numpy as np
from datetime import datetime
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ExportManager:
    def __init__(self):
        self.datetime_format = '%Y-%m-%d %H:%M:%S'
        self.date_format = '%Y-%m-%d'
    
    def export_to_excel(self, schedule, financial_analysis, quality_metrics, doctor_workload, cabinet_utilization):
        """Export comprehensive schedule report to Excel format"""
        
        # Create Excel workbook in memory
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Schedule worksheet
            schedule_df = self._prepare_schedule_for_export(schedule)
            schedule_df.to_excel(writer, sheet_name='Расписание', index=False)
            
            # Doctor workload worksheet
            if not doctor_workload.empty:
                doctor_workload.to_excel(writer, sheet_name='Загрузка врачей', index=False)
            
            # Cabinet utilization worksheet
            if not cabinet_utilization.empty:
                cabinet_utilization.to_excel(writer, sheet_name='Использование кабинетов', index=False)
            
            # Financial analysis worksheet
            financial_df = pd.DataFrame([financial_analysis])
            financial_df.to_excel(writer, sheet_name='Финансовый анализ', index=False)
            
            # Quality metrics worksheet
            quality_df = pd.DataFrame([quality_metrics])
            quality_df.to_excel(writer, sheet_name='Показатели качества', index=False)
            
            # Summary worksheet
            summary_df = self._create_summary_report(
                schedule, financial_analysis, quality_metrics, doctor_workload, cabinet_utilization
            )
            summary_df.to_excel(writer, sheet_name='Сводка', index=False)
            
            # Format worksheets
            self._format_excel_worksheets(writer)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def export_to_csv(self, schedule):
        """Export schedule to CSV format"""
        
        if not schedule:
            return "Дата,Врач ID,Кабинет ID,Смена,Время начала,Время окончания,Услуга,ДМС\n"
        
        schedule_df = self._prepare_schedule_for_export(schedule)
        
        # Convert to CSV
        buffer = io.StringIO()
        schedule_df.to_csv(buffer, index=False, encoding='utf-8')
        
        return buffer.getvalue()
    
    def export_analytics(self, financial_analysis, quality_metrics, evolution_history):
        """Export analytics data to JSON format"""
        
        analytics_data = {
            'export_timestamp': datetime.now().isoformat(),
            'financial_analysis': financial_analysis,
            'quality_metrics': quality_metrics,
            'evolution_history': evolution_history,
            'summary': {
                'total_revenue': financial_analysis.get('total_revenue', 0),
                'demand_coverage': financial_analysis.get('demand_coverage', 0),
                'avg_utilization': financial_analysis.get('avg_doctor_utilization', 0),
                'reliability_score': quality_metrics.get('reliability_score', 0),
                'strategy_score': quality_metrics.get('strategy_score', 0),
                'personnel_balance': quality_metrics.get('personnel_balance', 0)
            }
        }
        
        # Convert to JSON
        return json.dumps(analytics_data, ensure_ascii=False, indent=2, default=str)
    
    def _prepare_schedule_for_export(self, schedule):
        """Prepare schedule data for export"""
        
        if not schedule:
            return pd.DataFrame(columns=[
                'Дата', 'Врач ID', 'Кабинет ID', 'Смена', 
                'Время начала', 'Время окончания', 'Услуга', 'ДМС'
            ])
        
        schedule_df = pd.DataFrame(schedule)
        
        # Prepare export columns
        export_data = []
        
        for _, row in schedule_df.iterrows():
            export_data.append({
                'Дата': row['day'].strftime(self.date_format) if pd.notnull(row['day']) else '',
                'Врач ID': row.get('doctor_id', ''),
                'Кабинет ID': row.get('cabinet_id', ''),
                'Смена': self._translate_shift(row.get('shift', '')),
                'Время начала': row.get('start_time', ''),
                'Время окончания': row.get('end_time', ''),
                'Услуга': row.get('service', 'Не указана'),
                'ДМС': 'Да' if row.get('is_dms', False) else 'Нет'
            })
        
        return pd.DataFrame(export_data)
    
    def _translate_shift(self, shift):
        """Translate shift types to Russian"""
        
        translations = {
            'morning': 'Утро',
            'evening': 'Вечер',
            'day': 'День',
            'night': 'Ночь'
        }
        
        return translations.get(shift, shift)
    
    def _create_summary_report(self, schedule, financial_analysis, quality_metrics, doctor_workload, cabinet_utilization):
        """Create summary report"""
        
        summary_data = []
        
        # Basic statistics
        summary_data.append({
            'Показатель': 'Общее количество записей',
            'Значение': len(schedule) if schedule else 0,
            'Комментарий': 'Всего слотов в расписании'
        })
        
        summary_data.append({
            'Показатель': 'Прогнозируемая выручка (руб.)',
            'Значение': f"{financial_analysis.get('total_revenue', 0):,.0f}",
            'Комментарий': 'Ожидаемая месячная выручка'
        })
        
        summary_data.append({
            'Показатель': 'Покрытие спроса (%)',
            'Значение': f"{financial_analysis.get('demand_coverage', 0)*100:.1f}%",
            'Комментарий': 'Процент покрытия прогнозируемого спроса'
        })
        
        summary_data.append({
            'Показатель': 'Средняя загрузка врачей (%)',
            'Значение': f"{financial_analysis.get('avg_doctor_utilization', 0)*100:.1f}%",
            'Комментарий': 'Средний коэффициент загрузки'
        })
        
        summary_data.append({
            'Показатель': 'Коэффициент надежности',
            'Значение': f"{quality_metrics.get('reliability_score', 0):.3f}",
            'Комментарий': 'Оценка надежности врачей'
        })
        
        summary_data.append({
            'Показатель': 'Стратегический балл',
            'Значение': f"{quality_metrics.get('strategy_score', 0):.3f}",
            'Комментарий': 'Соответствие стратегическим целям'
        })
        
        summary_data.append({
            'Показатель': 'Баланс персонала',
            'Значение': f"{quality_metrics.get('personnel_balance', 0):.3f}",
            'Комментарий': 'Равномерность распределения нагрузки'
        })
        
        # Doctor statistics
        if not doctor_workload.empty:
            summary_data.append({
                'Показатель': 'Количество врачей',
                'Значение': len(doctor_workload),
                'Комментарий': 'Врачей в расписании'
            })
            
            summary_data.append({
                'Показатель': 'Максимальная нагрузка врача (ч)',
                'Значение': doctor_workload['total_hours'].max(),
                'Комментарий': 'Наиболее загруженный врач'
            })
            
            summary_data.append({
                'Показатель': 'Минимальная нагрузка врача (ч)',
                'Значение': doctor_workload['total_hours'].min(),
                'Комментарий': 'Наименее загруженный врач'
            })
        
        # Cabinet statistics
        if not cabinet_utilization.empty:
            avg_utilization = cabinet_utilization.groupby('cabinet_id')['utilization_rate'].mean()
            
            summary_data.append({
                'Показатель': 'Количество кабинетов',
                'Значение': cabinet_utilization['cabinet_id'].nunique(),
                'Комментарий': 'Кабинетов в использовании'
            })
            
            summary_data.append({
                'Показатель': 'Средняя загрузка кабинетов (%)',
                'Значение': f"{avg_utilization.mean():.1f}%",
                'Комментарий': 'Средняя загрузка по всем кабинетам'
            })
        
        return pd.DataFrame(summary_data)
    
    def _format_excel_worksheets(self, writer):
        """Apply formatting to Excel worksheets"""
        
        workbook = writer.book
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Apply header formatting
            for cell in worksheet[1]:  # First row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply borders to all cells with data
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = border
    
    def export_demand_forecast(self, demand_forecast):
        """Export demand forecast to CSV"""
        
        if demand_forecast.empty:
            return "Дата,Услуга,Прогноз спроса,ДМС спрос,Базовый прогноз,Коэф. сезонности,Коэф. акций,Коэф. буфера\n"
        
        # Prepare export data
        export_df = demand_forecast.copy()
        export_df['date'] = pd.to_datetime(export_df['date']).dt.strftime(self.date_format)
        
        # Rename columns to Russian
        column_mapping = {
            'date': 'Дата',
            'service': 'Услуга',
            'predicted_demand': 'Прогноз спроса',
            'dms_demand': 'ДМС спрос',
            'base_forecast': 'Базовый прогноз',
            'seasonal_factor': 'Коэф. сезонности',
            'promo_factor': 'Коэф. акций',
            'buffer_factor': 'Коэф. буфера'
        }
        
        export_df = export_df.rename(columns=column_mapping)
        
        # Select columns that exist
        available_columns = [col for col in column_mapping.values() if col in export_df.columns]
        export_df = export_df[available_columns]
        
        # Convert to CSV
        buffer = io.StringIO()
        export_df.to_csv(buffer, index=False, encoding='utf-8')
        
        return buffer.getvalue()
    
    def export_doctor_performance(self, financial_metrics, doctors_df):
        """Export doctor performance metrics to Excel"""
        
        if financial_metrics.empty or doctors_df.empty:
            buffer = io.BytesIO()
            empty_df = pd.DataFrame(columns=['Врач', 'Специальность', 'Выручка', 'Записей', 'Надежность'])
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                empty_df.to_excel(writer, sheet_name='Показатели врачей', index=False)
            buffer.seek(0)
            return buffer.getvalue()
        
        # Merge financial metrics with doctor information
        performance_df = financial_metrics.merge(
            doctors_df[['doctor_id', 'name', 'specialty', 'experience_years']],
            on='doctor_id',
            how='left'
        )
        
        # Prepare export columns
        export_columns = {
            'name': 'Врач',
            'specialty': 'Специальность',
            'experience_years': 'Стаж (лет)',
            'total_revenue': 'Общая выручка (руб.)',
            'total_appointments': 'Всего записей',
            'avg_appointment_cost': 'Средняя стоимость приема (руб.)',
            'reliability_coefficient': 'Коэффициент надежности',
            'fill_rate': 'Коэффициент заполняемости',
            'service_diversity': 'Разнообразие услуг',
            'dms_ratio': 'Доля ДМС',
            'months_active': 'Месяцев активности'
        }
        
        # Select and rename columns
        available_columns = [col for col in export_columns.keys() if col in performance_df.columns]
        export_df = performance_df[available_columns].copy()
        export_df = export_df.rename(columns={k: v for k, v in export_columns.items() if k in available_columns})
        
        # Format numeric columns
        numeric_columns = ['Общая выручка (руб.)', 'Средняя стоимость приема (руб.)']
        for col in numeric_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].round(2)
        
        ratio_columns = ['Коэффициент надежности', 'Коэффициент заполняемости', 'Доля ДМС']
        for col in ratio_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].round(3)
        
        # Export to Excel
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name='Показатели врачей', index=False)
            self._format_excel_worksheets(writer)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_optimization_report(self, schedule, evolution_history, fitness_breakdown):
        """Create comprehensive optimization report"""
        
        report_data = {
            'optimization_summary': {
                'timestamp': datetime.now().isoformat(),
                'total_assignments': len(schedule) if schedule else 0,
                'optimization_generations': len(evolution_history) if evolution_history else 0,
                'final_fitness': evolution_history[-1]['best_fitness'] if evolution_history else 0,
                'improvement': (evolution_history[-1]['best_fitness'] - evolution_history[0]['best_fitness']) if len(evolution_history) > 1 else 0
            },
            'fitness_components': fitness_breakdown,
            'evolution_statistics': {
                'best_fitness_trend': [gen['best_fitness'] for gen in evolution_history] if evolution_history else [],
                'avg_fitness_trend': [gen['avg_fitness'] for gen in evolution_history] if evolution_history else [],
                'convergence_analysis': self._analyze_convergence(evolution_history)
            },
            'schedule_statistics': self._calculate_schedule_statistics(schedule)
        }
        
        return json.dumps(report_data, ensure_ascii=False, indent=2, default=str)
    
    def _analyze_convergence(self, evolution_history):
        """Analyze convergence patterns in evolution"""
        
        if not evolution_history or len(evolution_history) < 5:
            return {'status': 'insufficient_data'}
        
        best_fitnesses = [gen['best_fitness'] for gen in evolution_history]
        
        # Calculate improvement rate over last 10 generations
        last_10_generations = best_fitnesses[-10:]
        improvement_rate = (last_10_generations[-1] - last_10_generations[0]) / len(last_10_generations) if len(last_10_generations) > 1 else 0
        
        # Detect plateau
        plateau_threshold = 0.001
        plateau_generations = 0
        
        for i in range(len(best_fitnesses) - 1, 0, -1):
            if abs(best_fitnesses[i] - best_fitnesses[i-1]) < plateau_threshold:
                plateau_generations += 1
            else:
                break
        
        convergence_status = 'converged' if plateau_generations >= 5 else 'improving' if improvement_rate > plateau_threshold else 'slow_progress'
        
        return {
            'status': convergence_status,
            'plateau_generations': plateau_generations,
            'improvement_rate': improvement_rate,
            'total_improvement': best_fitnesses[-1] - best_fitnesses[0] if len(best_fitnesses) > 1 else 0
        }
    
    def _calculate_schedule_statistics(self, schedule):
        """Calculate detailed schedule statistics"""
        
        if not schedule:
            return {'total_assignments': 0}
        
        schedule_df = pd.DataFrame(schedule)
        
        statistics = {
            'total_assignments': len(schedule_df),
            'unique_doctors': schedule_df['doctor_id'].nunique(),
            'unique_cabinets': schedule_df['cabinet_id'].nunique(),
            'unique_days': schedule_df['day'].nunique() if 'day' in schedule_df.columns else 0,
            'shift_distribution': schedule_df['shift'].value_counts().to_dict() if 'shift' in schedule_df.columns else {},
            'dms_percentage': schedule_df.get('is_dms', pd.Series([False] * len(schedule_df))).mean() * 100
        }
        
        # Service distribution
        if 'service' in schedule_df.columns:
            statistics['service_distribution'] = schedule_df['service'].value_counts().to_dict()
        
        # Time distribution
        if 'start_time' in schedule_df.columns:
            schedule_df['hour'] = schedule_df['start_time'].str[:2].astype(int, errors='ignore')
            statistics['hourly_distribution'] = schedule_df['hour'].value_counts().to_dict()
        
        return statistics

